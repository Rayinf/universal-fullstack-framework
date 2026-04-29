from __future__ import annotations

import io
from typing import Any

from fastapi.responses import JSONResponse, StreamingResponse

try:
  import openpyxl  # type: ignore
except ImportError:
  openpyxl = None  # type: ignore


def export_to_excel(
  headers: list[str],
  rows: list[list[Any]],
  sheet_name: str = 'Sheet1',
) -> StreamingResponse | JSONResponse:
  if openpyxl is None:
    return JSONResponse(
      status_code=500,
      content={'code': 500, 'msg': '服务端未安装 openpyxl，无法导出 Excel'},
    )

  wb = openpyxl.Workbook()
  ws = wb.active
  ws.title = sheet_name

  for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = openpyxl.styles.Font(bold=True)

  for row_idx, row_data in enumerate(rows, 2):
    for col_idx, value in enumerate(row_data, 1):
      ws.cell(row=row_idx, column=col_idx, value=value)

  for col_idx, header in enumerate(headers, 1):
    max_len = len(str(header))
    for row_data in rows[:100]:
      if col_idx - 1 < len(row_data):
        max_len = max(max_len, len(str(row_data[col_idx - 1] or '')))
    ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 50)

  buffer = io.BytesIO()
  wb.save(buffer)
  buffer.seek(0)

  from urllib.parse import quote

  encoded_name = quote(f'{sheet_name}.xlsx')
  return StreamingResponse(
    buffer,
    media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    headers={'Content-Disposition': f"attachment; filename*=UTF-8''{encoded_name}"},
  )
